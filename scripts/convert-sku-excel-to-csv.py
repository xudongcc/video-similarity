from __future__ import annotations

import csv
import re
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
SKU_DIR = ROOT / "sku"
INPUT_GLOB = "*.xlsx"
MD5_HEADER_CANDIDATES = {
    "md5",
    "material_md5",
    "video_md5",
    "link_md5",
}


def normalize_header(value: object) -> str:
    text = "" if value is None else str(value)
    return text.strip().lower().replace(" ", "_")


def parse_filename_parts(path: Path) -> tuple[str, int | None]:
    stem = path.stem
    match = re.fullmatch(r"([A-Za-z]+[0-9]+)(?:_(\d+))?", stem)
    if not match:
        return stem, None
    sku = match.group(1)
    expected_count = int(match.group(2)) if match.group(2) else None
    return sku, expected_count


def find_md5_column(header_row: tuple[object, ...]) -> int:
    normalized = [normalize_header(value) for value in header_row]
    for index, name in enumerate(normalized):
        if name in MD5_HEADER_CANDIDATES:
            return index
    raise ValueError(f"Unable to find md5 column in header: {header_row!r}")


def load_md5_values(path: Path) -> list[str]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook[workbook.sheetnames[0]]
    rows = sheet.iter_rows(values_only=True)
    header = next(rows, None)
    if header is None:
        return []

    md5_index = find_md5_column(header)
    values: list[str] = []
    seen: set[str] = set()

    for row in rows:
        if row is None:
            continue
        if md5_index >= len(row):
            continue
        raw_value = row[md5_index]
        md5 = "" if raw_value is None else str(raw_value).strip()
        if not md5 or md5 in seen:
            continue
        seen.add(md5)
        values.append(md5)

    return values


def write_csv(path: Path, sku: str, md5_values: list[str]) -> None:
    with path.open("w", newline="", encoding="utf-8-sig") as file:
        writer = csv.writer(file)
        writer.writerow(["md5", "sku"])
        for md5 in md5_values:
            writer.writerow([md5, sku])


def main() -> None:
    excel_files = sorted(SKU_DIR.glob(INPUT_GLOB))
    if not excel_files:
        raise SystemExit(f"No Excel files found in {SKU_DIR}")

    reports: list[dict[str, object]] = []

    for excel_path in excel_files:
        sku, expected_count = parse_filename_parts(excel_path)
        csv_path = excel_path.with_name(f"{sku}.csv")
        md5_values = load_md5_values(excel_path)
        write_csv(csv_path, sku, md5_values)
        reports.append(
            {
                "excel": excel_path.name,
                "csv": csv_path.name,
                "sku": sku,
                "expected_count": expected_count,
                "actual_count": len(md5_values),
                "count_matches_filename": expected_count == len(md5_values)
                if expected_count is not None
                else None,
            }
        )

    for item in reports:
        print(
            "\t".join(
                [
                    str(item["excel"]),
                    str(item["csv"]),
                    str(item["sku"]),
                    str(item["actual_count"]),
                    str(item["expected_count"]),
                    str(item["count_matches_filename"]),
                ]
            )
        )


if __name__ == "__main__":
    main()
